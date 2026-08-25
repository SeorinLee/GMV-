"""Regenerate the static upload template at apps/web/public/gmv_upload_template.xlsx (spec §6).

Run: ``python tools/create_excel_template.py``. The template has ONLY a ``Creator Name`` column —
GMV and 판매개수 are added automatically to the result file, so users never create them.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
OUT = (ROOT / ".." / "apps" / "web" / "public" / "gmv_upload_template.xlsx").resolve()

LAST_INPUT_ROW = 31  # A2:A31 -> 30 empty input rows

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")   # dark blue
INPUT_FILL = PatternFill("solid", fgColor="FFF7E6")    # light amber (user input)
HEADER_FONT = Font(bold=True, color="FFFFFF")


def create_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Creators"

    header = ws.cell(row=1, column=1, value="Creator Name")
    header.fill = HEADER_FILL
    header.font = HEADER_FONT
    header.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx in range(2, LAST_INPUT_ROW + 1):
        ws.cell(row=row_idx, column=1).fill = INPUT_FILL

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 32
    ws.row_dimensions[1].height = 22

    info = wb.create_sheet("사용방법")
    info.column_dimensions["A"].width = 80
    lines = [
        "1. Creator Name 열에 TikTok creator username만 입력하세요.",
        "2. GMV 조회 시작을 누르면 결과 파일에 GMV와 판매개수가 자동으로 추가됩니다.",
        "3. GMV와 판매개수 컬럼은 사용자가 직접 만들 필요 없습니다.",
    ]
    for i, line in enumerate(lines, start=1):
        info.cell(row=i, column=1, value=line)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    create_template(OUT)
    print(f"Created {OUT}")
