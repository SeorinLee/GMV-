"""Invitation Acceptor parser, safety, persistence and control tests."""

import asyncio
import io

from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook

from gmv.api import create_app
from gmv.automation.invitation_inspector_session import (
    INVITATION_SEARCH_TABS,
    CreatorDetails,
    InvitationInspection,
    InvitationNotFoundError,
    creator_activity_flags,
    inspect_json_for_creators,
    invitation_suffix_query,
    invitation_tab_pattern,
    merge_creator_pages,
    target_url_from_authenticated_page,
)
from gmv.config import ProfileStatus, get_profile
from gmv.invitation_acceptor import (
    expand_invitation_input,
    group_orders_by_product,
    is_exact_invitation_match,
    parse_invitation_name,
)
from gmv.invitation_acceptor_service import creator_is_member
from gmv.store import JobStore


def test_right_based_parser_range_order_and_product_grouping():
    parsed = parse_invitation_name("OWNER_WITH_UNDERSCORE_SZP_0810_15")
    assert parsed.owner == "OWNER_WITH_UNDERSCORE"
    assert (parsed.product, parsed.date, parsed.number) == ("SZP", "0810", "15")
    specs, errors = expand_invitation_input(
        "PJH_SZP_0810_1~3, D_MC_0720_15\nD_KTB_0430_1~2"
    )
    assert errors == []
    assert [spec.full_name for spec in specs] == [
        "PJH_SZP_0810_1",
        "PJH_SZP_0810_2",
        "PJH_SZP_0810_3",
        "D_MC_0720_15",
        "D_KTB_0430_1",
        "D_KTB_0430_2",
    ]
    assert group_orders_by_product(specs) == {
        "szp": [1, 2, 3],
        "mc": [4],
        "ktb": [5, 6],
    }


def test_exact_match_never_accepts_numeric_suffix_or_owner_variants():
    target = "PJH_SZP_0810_1"
    assert is_exact_invitation_match("pjh_szp_0810_1", target)
    assert not is_exact_invitation_match("PJH_SZP_0810_10", target)
    assert not is_exact_invitation_match("PJH_SZP_0810_11", target)
    assert not is_exact_invitation_match("ABC_SZP_0810_1", target)


def test_invitation_suffix_query_respects_tiktok_ten_character_limit():
    assert INVITATION_SEARCH_TABS == ("Ongoing",)
    assert invitation_suffix_query("PJH_SZPEUKA_0814_59") == "A_0814_59"
    assert invitation_suffix_query("PJH_SZPEUKA_0814_55") == "A_0814_55"
    assert invitation_suffix_query("SHORT") == "SHORT"
    assert invitation_tab_pattern("Ongoing").fullmatch("Ongoing 1")
    assert invitation_tab_pattern("Ongoing").fullmatch("Ongoing")
    assert not invitation_tab_pattern("Ongoing").fullmatch("Completed 1")


def test_target_route_preserves_authenticated_shop_parameters():
    url = target_url_from_authenticated_page(
        get_profile("US_EDGE"),
        "https://seller-us.tiktok.com/affiliate/landing?shop_region=US",
        "https://affiliate-us.tiktok.com/affiliate/collaboration/target-invitation?shop_region=US&shop_id=7495830785034323995&route_migration=1&tab=1",
    )
    assert url == (
        "https://affiliate-us.tiktok.com/affiliate/collaboration/target-invitation"
        "?shop_region=US&shop_id=7495830785034323995&route_migration=1&tab=1"
    )


def test_creator_pages_preserve_invited_order_and_dedupe_by_id():
    merged = merge_creator_pages(
        [
            [CreatorDetails("creator_a", "A", "10000001", "US")],
            [
                CreatorDetails("renamed_a", "A2", "10000001", "US"),
                CreatorDetails("creator_b", "B", "10000002", "US"),
            ],
        ]
    )
    assert [item.creator for item in merged] == ["creator_a", "creator_b"]


def test_creator_network_json_extracts_id_nickname_and_region():
    creators = inspect_json_for_creators(
        {"data": {"list": [{"unique_id": "milannlynn", "nickname": "Milan", "creator_id": "74954847", "region": "US"}]}}
    )
    assert creators == [CreatorDetails("milannlynn", "Milan", "74954847", "US")]


def test_invited_creator_activity_columns_become_exact_o_flags():
    assert creator_activity_flags(["1 product", "0 products", "View details"]) == (True, False)
    assert creator_activity_flags(["2 products", "3 videos"]) == (True, True)
    assert creator_activity_flags(["0 products", "0 contents"]) == (False, False)


def test_unique_nickname_fallback_matches_when_tab_hides_creator_id():
    invited = [
        CreatorDetails("creator_a", "Unique Nick", "74950001", "US"),
        CreatorDetails("creator_b", "Other", "74950002", "US"),
    ]
    added = [CreatorDetails("masked_name", "Unique Nick", "", "US")]
    assert creator_is_member(invited[0], added, invited, set())
    assert not creator_is_member(invited[1], added, invited, set())


class FakeAcceptorSession:
    calls = []

    def __init__(self, profile):
        self.profile = profile
        self.cancel_event = None

    async def start(self):
        return ProfileStatus.CONNECTED

    async def inspect_invitation(self, spec, progress_cb=None):
        type(self).calls.append((spec.full_name, self.profile.storage_root))
        if progress_cb:
            await progress_cb("search", f"{spec.product} search", 1, 2)
        creators = [
            CreatorDetails("creator_a", "Creator A", "74950001", "US"),
            CreatorDetails("creator_b", "Creator B", "74950002", "US"),
        ]
        return InvitationInspection(
            creators=creators,
            added_product_keys={creators[0].identity},
            posted_content_keys={creators[1].identity},
        )

    async def close(self):
        return None


async def poll(client, job_id, terminal, tries=400):
    for _ in range(tries):
        payload = (await client.get(f"/invitation-accept-jobs/{job_id}")).json()
        if payload["status"] in terminal:
            return payload
        await asyncio.sleep(0.01)
    raise AssertionError(f"job did not reach {terminal}")


def test_acceptor_api_persists_order_uses_one_clone_and_exports_two_sheets(tmp_path, monkeypatch):
    async def go():
        FakeAcceptorSession.calls.clear()
        monkeypatch.setenv("GMV_PROFILE_ROOT", str(tmp_path / "profiles"))
        monkeypatch.setenv("GMV_RUNTIME_PROFILE_ROOT", str(tmp_path / "runtime"))
        app = create_app(
            JobStore(str(tmp_path / "jobs")),
            invitation_acceptor_session_factory=FakeAcceptorSession,
        )
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            parsed = await client.post(
                "/invitation-accept/parse",
                json={"profile_code": "US_EDGE", "invitation_text": "PJH_SZP_0810_1~2"},
            )
            assert [item["number"] for item in parsed.json()["items"]] == ["1", "2"]
            created = await client.post(
                "/invitation-accept-jobs",
                json={"profile_code": "US_EDGE", "invitation_text": "PJH_SZP_0810_1~2"},
            )
            assert created.status_code == 200
            payload = await poll(client, created.json()["job_id"], {"completed"})
            assert payload["concurrency"] == 1
            assert [state["status"] for state in payload["invitation_accept_states"]] == [
                "SUCCESS",
                "SUCCESS",
            ]
            assert len(payload["invitation_creator_rows"]) == 4
            assert [call[0] for call in FakeAcceptorSession.calls] == [
                "PJH_SZP_0810_1",
                "PJH_SZP_0810_2",
            ]
            assert all("runtime" in call[1] for call in FakeAcceptorSession.calls)
            response = await client.get(
                f"/invitation-accept-jobs/{created.json()['job_id']}/download"
            )
            workbook = load_workbook(io.BytesIO(response.content))
            assert workbook.sheetnames == ["Results", "Errors"]
            assert [workbook["Results"].cell(row, 1).value for row in (2, 3)] == ["SZP", "SZP"]
            assert workbook["Results"]["G2"].value == "O"
            assert workbook["Results"]["H3"].value == "O"
            assert workbook["Errors"].max_row == 1

    asyncio.run(go())


def test_invalid_input_is_rejected_before_browser_start(tmp_path):
    async def go():
        app = create_app(
            JobStore(str(tmp_path / "jobs")),
            invitation_acceptor_session_factory=FakeAcceptorSession,
        )
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            response = await client.post(
                "/invitation-accept-jobs",
                json={"profile_code": "US_CHROME", "invitation_text": "BAD_NAME"},
            )
            assert response.status_code == 400
            assert response.json()["code"] == "INVALID_INVITATION_INPUT"

    asyncio.run(go())


def test_pause_then_resume_continues_next_invitation(tmp_path, monkeypatch):
    async def go():
        entered = asyncio.Event()
        release = asyncio.Event()

        class GateSession(FakeAcceptorSession):
            async def inspect_invitation(self, spec, progress_cb=None):
                if spec.number == "1":
                    entered.set()
                    await release.wait()
                return await super().inspect_invitation(spec, progress_cb)

        monkeypatch.setenv("GMV_PROFILE_ROOT", str(tmp_path / "profiles"))
        monkeypatch.setenv("GMV_RUNTIME_PROFILE_ROOT", str(tmp_path / "runtime"))
        app = create_app(
            JobStore(str(tmp_path / "jobs")),
            invitation_acceptor_session_factory=GateSession,
        )
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            created = await client.post(
                "/invitation-accept-jobs",
                json={"profile_code": "US_CHROME", "invitation_text": "PJH_SZP_0810_1~2"},
            )
            job_id = created.json()["job_id"]
            await asyncio.wait_for(entered.wait(), timeout=3)
            assert (await client.post(f"/invitation-accept-jobs/{job_id}/pause")).status_code == 200
            release.set()
            for _ in range(200):
                paused = (await client.get(f"/invitation-accept-jobs/{job_id}")).json()
                if paused["status"] == "paused" and paused["processed"] == 1:
                    break
                await asyncio.sleep(0.01)
            else:
                raise AssertionError("job did not pause between invitations")
            assert (await client.post(f"/invitation-accept-jobs/{job_id}/resume")).status_code == 200
            completed = await poll(client, job_id, {"completed"})
            assert completed["processed"] == 2

    asyncio.run(go())


def test_cancel_keeps_checkpoint_and_marks_remaining_items(tmp_path, monkeypatch):
    async def go():
        entered = asyncio.Event()
        release = asyncio.Event()

        class GateSession(FakeAcceptorSession):
            async def inspect_invitation(self, spec, progress_cb=None):
                entered.set()
                await release.wait()
                return await super().inspect_invitation(spec, progress_cb)

        monkeypatch.setenv("GMV_PROFILE_ROOT", str(tmp_path / "profiles"))
        monkeypatch.setenv("GMV_RUNTIME_PROFILE_ROOT", str(tmp_path / "runtime"))
        app = create_app(
            JobStore(str(tmp_path / "jobs")),
            invitation_acceptor_session_factory=GateSession,
        )
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            created = await client.post(
                "/invitation-accept-jobs",
                json={"profile_code": "US_CHROME", "invitation_text": "PJH_SZP_0810_1~2"},
            )
            job_id = created.json()["job_id"]
            await asyncio.wait_for(entered.wait(), timeout=3)
            assert (await client.post(f"/invitation-accept-jobs/{job_id}/cancel")).status_code == 200
            release.set()
            cancelled = await poll(client, job_id, {"cancelled"})
            assert cancelled["processed"] == 2
            assert [item["status"] for item in cancelled["invitation_accept_states"]] == [
                "SUCCESS",
                "CANCELLED",
            ]
            assert (await client.get(f"/invitation-accept-jobs/{job_id}/download")).status_code == 200

    asyncio.run(go())


def test_retry_runs_only_failed_items(tmp_path, monkeypatch):
    async def go():
        calls = []

        class FlakySession(FakeAcceptorSession):
            async def inspect_invitation(self, spec, progress_cb=None):
                calls.append(spec.full_name)
                if spec.number == "1" and calls.count(spec.full_name) == 1:
                    raise InvitationNotFoundError("not found on first run")
                return await super().inspect_invitation(spec, progress_cb)

        monkeypatch.setenv("GMV_PROFILE_ROOT", str(tmp_path / "profiles"))
        monkeypatch.setenv("GMV_RUNTIME_PROFILE_ROOT", str(tmp_path / "runtime"))
        app = create_app(
            JobStore(str(tmp_path / "jobs")),
            invitation_acceptor_session_factory=FlakySession,
        )
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            created = await client.post(
                "/invitation-accept-jobs",
                json={"profile_code": "US_EDGE", "invitation_text": "PJH_SZP_0810_1~2"},
            )
            job_id = created.json()["job_id"]
            first = await poll(client, job_id, {"completed_with_errors"})
            assert [item["status"] for item in first["invitation_accept_states"]] == [
                "NOT_FOUND",
                "SUCCESS",
            ]
            assert (await client.post(f"/invitation-accept-jobs/{job_id}/retry")).status_code == 200
            final = await poll(client, job_id, {"completed"})
            assert [item["status"] for item in final["invitation_accept_states"]] == [
                "SUCCESS",
                "SUCCESS",
            ]
            assert calls == ["PJH_SZP_0810_1", "PJH_SZP_0810_2", "PJH_SZP_0810_1"]

    asyncio.run(go())
