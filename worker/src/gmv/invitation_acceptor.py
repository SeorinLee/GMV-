"""Invitation Acceptor parsing, safe identity matching and XLSX export."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

_INVITATION_RE = re.compile(r"^(.*?)_([^_]+)_([^_]+)_(\d+)$")
_RANGE_RE = re.compile(r"^(.*?)_(\d+)~(\d+)$")


class InvitationInputError(ValueError):
    def __init__(self, detail: str, code: str = "INVALID_INVITATION_INPUT"):
        super().__init__(detail)
        self.detail = detail
        self.code = code


@dataclass(frozen=True)
class InvitationSpec:
    order: int
    full_name: str
    owner: str
    product: str
    date: str
    number: str


def normalize_invitation_name(value: object) -> str:
    """Normalize only whitespace and case; punctuation and numeric suffix stay significant."""
    return str(value or "").strip().casefold()


def parse_invitation_name(full_name: str, *, order: int = 1) -> InvitationSpec:
    name = str(full_name or "").strip()
    match = _INVITATION_RE.fullmatch(name)
    if not match or not match.group(1):
        raise InvitationInputError(
            f"형식을 확인해주세요: {name or '(빈 값)'}. 예: PJH_SZP_0810_1"
        )
    return InvitationSpec(
        order=order,
        full_name=name,
        owner=match.group(1),
        product=match.group(2),
        date=match.group(3),
        number=match.group(4),
    )


def expand_invitation_input(raw: str) -> tuple[list[InvitationSpec], list[str]]:
    """Expand final-number ranges and preserve the user's first-appearance order."""
    tokens = [part.strip() for part in re.split(r"[\r\n,]+", str(raw or "")) if part.strip()]
    expanded: list[str] = []
    errors: list[str] = []
    for token in tokens:
        range_match = _RANGE_RE.fullmatch(token)
        if range_match:
            start, end = int(range_match.group(2)), int(range_match.group(3))
            if start > end or end - start > 5000:
                errors.append(f"올바르지 않은 범위: {token}")
                continue
            expanded.extend(f"{range_match.group(1)}_{number}" for number in range(start, end + 1))
        else:
            expanded.append(token)

    specs: list[InvitationSpec] = []
    seen: set[str] = set()
    for full_name in expanded:
        key = normalize_invitation_name(full_name)
        if key in seen:
            continue
        try:
            spec = parse_invitation_name(full_name, order=len(specs) + 1)
        except InvitationInputError as exc:
            errors.append(exc.detail)
            continue
        seen.add(key)
        specs.append(spec)
    return specs, errors


def is_exact_invitation_match(actual: str, requested: str) -> bool:
    return normalize_invitation_name(actual) == normalize_invitation_name(requested)


def group_orders_by_product(specs: list[InvitationSpec]) -> dict[str, list[int]]:
    grouped: dict[str, list[int]] = {}
    for spec in specs:
        grouped.setdefault(spec.product.casefold(), []).append(spec.order)
    return grouped


def write_invitation_accept_results(states, output_path: str, creator_rows=None) -> str:
    workbook = Workbook()
    results = workbook.active
    results.title = "Results"
    results.append(
        [
            "Keyword",
            "Invitation",
            "Creator",
            "Nickname",
            "Creator ID",
            "Region",
            "Added products",
            "Posted content",
        ]
    )
    for row in sorted(creator_rows or [], key=lambda item: item.order):
        results.append(
            [
                row.keyword,
                row.invitation_name,
                row.creator,
                row.nickname,
                row.creator_id,
                row.region or row.market,
                "O" if row.added_products else None,
                "O" if row.posted_content else None,
            ]
        )

    errors = workbook.create_sheet("Errors")
    errors.append(["Order", "Invitation", "Error", "Time"])
    successful = {"SUCCESS"}
    for state in sorted(states, key=lambda item: item.order):
        if state.status in successful | {"QUEUED", "PROCESSING"}:
            continue
        errors.append([state.order, state.invitation_name, state.message or state.status, state.processed_at])

    for sheet in (results, errors):
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4968ED")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    for column, width in {"A": 16, "B": 32, "C": 24, "D": 24, "E": 20, "F": 12, "G": 18, "H": 18}.items():
        results.column_dimensions[column].width = width
    for column, width in {"A": 10, "B": 32, "C": 60, "D": 24}.items():
        errors.column_dimensions[column].width = width
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    return output_path
