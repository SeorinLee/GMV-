"""API-level proof that jobs sharing one login profile run from independent clones."""

import asyncio
import io
from pathlib import Path

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook

from gmv.api import create_app
from gmv.config import ProfileStatus, get_profile
from gmv.gmv_parser import parse_gmv
from gmv.models import ErrorCode, JobCancelledError, LookupResult, RowStatus
from gmv.store import JobStore


def _xlsx_bytes(name="alpha"):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Creator Name"
    ws["A2"] = name
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def _poll(client, job_id, terminal, tries=400):
    for _ in range(tries):
        payload = (await client.get(f"/jobs/{job_id}")).json()
        if payload["status"] in terminal:
            return payload
        await asyncio.sleep(0.01)
    raise AssertionError(f"job {job_id} did not reach {terminal}")


async def _wait_paths_removed(paths, tries=200):
    for _ in range(tries):
        if all(not path.exists() for path in paths):
            return
        await asyncio.sleep(0.005)
    raise AssertionError(f"runtime profiles were not cleaned: {paths}")


@pytest.fixture
def isolated_profile_roots(tmp_path, monkeypatch):
    profile_root = tmp_path / "persistent_profiles"
    runtime_root = tmp_path / "runtime_profiles"
    monkeypatch.setenv("GMV_PROFILE_ROOT", str(profile_root))
    monkeypatch.setenv("GMV_RUNTIME_PROFILE_ROOT", str(runtime_root))
    for code in ("US_CHROME", "UK_EDGE"):
        cookies = Path(get_profile(code).storage_root) / "Default" / "Cookies"
        cookies.parent.mkdir(parents=True, exist_ok=True)
        cookies.write_text(f"logged-in-{code}", encoding="utf-8")
    return runtime_root


def test_same_profile_three_jobs_run_concurrently_with_distinct_clones(
    tmp_path, isolated_profile_roots
):
    async def go():
        release = asyncio.Event()
        all_running = asyncio.Event()

        class GateSession:
            paths = []
            starts = []
            running = 0
            max_running = 0

            def __init__(self, profile):
                self.profile = profile
                self.cancel_event = None
                type(self).paths.append(Path(profile.storage_root))

            async def start(self, concurrency=1):
                type(self).starts.append(concurrency)
                assert Path(self.profile.storage_root, "Default", "Cookies").exists()
                return ProfileStatus.CONNECTED

            async def search_creator(self, name):
                type(self).running += 1
                type(self).max_running = max(type(self).max_running, type(self).running)
                if type(self).running == 3:
                    all_running.set()
                try:
                    await release.wait()
                    return LookupResult(
                        normalized_username=name,
                        gmv=parse_gmv("$5K"),
                        status=RowStatus.SUCCESS,
                    )
                finally:
                    type(self).running -= 1

            async def close(self):
                return None

        app = create_app(
            store=JobStore(str(tmp_path / "jobs")),
            session_factory=GateSession,
        )
        async with AsyncClient(
            transport=ASGITransport(app=app), base_url="http://test"
        ) as client:
            responses = await asyncio.gather(
                *(
                    client.post(
                        "/jobs",
                        data={"profile_code": "US_CHROME", "concurrency": "7"},
                        files={"file": (f"job-{index}.xlsx", _xlsx_bytes(), "application/xlsx")},
                    )
                    for index in range(3)
                )
            )
            assert all(response.status_code == 200 for response in responses)
            job_ids = [response.json()["job_id"] for response in responses]
            await asyncio.wait_for(all_running.wait(), timeout=3)

            assert GateSession.max_running == 3
            assert GateSession.starts == [1, 1, 1]
            assert len(set(GateSession.paths)) == 3
            assert all(path.exists() for path in GateSession.paths)
            assert all(
                app.state.store.get(job_id).status == "running" for job_id in job_ids
            )
            assert app.state.busy_profiles == set()

            release.set()
            await asyncio.gather(
                *(_poll(client, job_id, {"completed"}) for job_id in job_ids)
            )
            assert all(not path.exists() for path in GateSession.paths)

    asyncio.run(go())


def test_different_profiles_still_run_concurrently(tmp_path, isolated_profile_roots):
    async def go():
        release = asyncio.Event()
        both_running = asyncio.Event()

        class GateSession:
            running = set()

            def __init__(self, profile):
                self.profile = profile

            async def start(self, concurrency=1):
                return ProfileStatus.CONNECTED

            async def search_creator(self, name):
                type(self).running.add(self.profile.profile_code)
                if len(type(self).running) == 2:
                    both_running.set()
                await release.wait()
                return LookupResult(
                    normalized_username=name,
                    gmv=parse_gmv("$5K"),
                    status=RowStatus.SUCCESS,
                )

            async def close(self):
                return None

        app = create_app(JobStore(str(tmp_path / "jobs")), GateSession)
        async with AsyncClient(
            transport=ASGITransport(app=app), base_url="http://test"
        ) as client:
            responses = await asyncio.gather(
                *(
                    client.post(
                        "/jobs",
                        data={"profile_code": code, "concurrency": "1"},
                        files={"file": (f"{code}.xlsx", _xlsx_bytes(), "application/xlsx")},
                    )
                    for code in ("US_CHROME", "UK_EDGE")
                )
            )
            await asyncio.wait_for(both_running.wait(), timeout=3)
            release.set()
            await asyncio.gather(
                *(
                    _poll(client, response.json()["job_id"], {"completed"})
                    for response in responses
                )
            )

    asyncio.run(go())


def test_job_waits_for_source_mutation_only_during_snapshot(
    tmp_path, isolated_profile_roots
):
    async def go():
        class ImmediateSession:
            created = 0

            def __init__(self, _profile):
                type(self).created += 1

            async def start(self, concurrency=1):
                return ProfileStatus.CONNECTED

            async def search_creator(self, name):
                return LookupResult(
                    normalized_username=name,
                    gmv=parse_gmv("$5K"),
                    status=RowStatus.SUCCESS,
                )

            async def close(self):
                return None

        app = create_app(JobStore(str(tmp_path / "jobs")), ImmediateSession)
        source_lock = app.state.profile_source_locks["US_CHROME"]
        await source_lock.acquire()
        async with AsyncClient(
            transport=ASGITransport(app=app), base_url="http://test"
        ) as client:
            response = await client.post(
                "/jobs",
                files={"file": ("waiting.xlsx", _xlsx_bytes(), "application/xlsx")},
            )
            assert response.status_code == 200
            await asyncio.sleep(0.03)
            assert ImmediateSession.created == 0
            assert app.state.store.get(response.json()["job_id"]).status == "queued"

            source_lock.release()
            await _poll(client, response.json()["job_id"], {"completed"})
            assert ImmediateSession.created == 1

    asyncio.run(go())


def test_cancel_and_browser_start_failure_cleanup_clones(tmp_path, isolated_profile_roots):
    async def go():
        entered = asyncio.Event()

        class CancelSession:
            paths = []

            def __init__(self, profile):
                self.profile = profile
                self.cancel_event = None
                type(self).paths.append(Path(profile.storage_root))

            async def start(self, concurrency=1):
                return ProfileStatus.CONNECTED

            async def search_creator(self, _name):
                entered.set()
                while self.cancel_event is None or not self.cancel_event.is_set():
                    await asyncio.sleep(0.005)
                raise JobCancelledError

            async def close(self):
                return None

        store = JobStore(str(tmp_path / "cancel-jobs"))
        app = create_app(store, CancelSession)
        async with AsyncClient(
            transport=ASGITransport(app=app), base_url="http://test"
        ) as client:
            response = await client.post(
                "/jobs",
                files={"file": ("cancel.xlsx", _xlsx_bytes(), "application/xlsx")},
            )
            job_id = response.json()["job_id"]
            await entered.wait()
            assert (await client.post(f"/jobs/{job_id}/cancel")).status_code == 200
            result = await _poll(client, job_id, {"cancelled"})
            assert result["output_filename"] is not None
            await _wait_paths_removed(CancelSession.paths)

        class StartFailureSession:
            paths = []

            def __init__(self, profile):
                self.profile = profile
                type(self).paths.append(Path(profile.storage_root))

            async def start(self, concurrency=1):
                raise RuntimeError("browser launch failed")

            async def close(self):
                return None

        failed_app = create_app(
            JobStore(str(tmp_path / "failed-jobs")), StartFailureSession
        )
        async with AsyncClient(
            transport=ASGITransport(app=failed_app), base_url="http://test"
        ) as client:
            response = await client.post(
                "/jobs",
                files={"file": ("failed.xlsx", _xlsx_bytes(), "application/xlsx")},
            )
            result = await _poll(
                client,
                response.json()["job_id"],
                {"completed_with_errors", "failed"},
            )
            assert result["status"] == "completed_with_errors"
            await _wait_paths_removed(StartFailureSession.paths)

    asyncio.run(go())


def test_retry_gets_new_clone_and_watchdog_restart_reuses_job_clone(
    tmp_path, isolated_profile_roots, monkeypatch
):
    async def go():
        class RetrySession:
            paths = []
            succeed = False

            def __init__(self, profile):
                self.profile = profile
                type(self).paths.append(Path(profile.storage_root))

            async def start(self, concurrency=1):
                return ProfileStatus.CONNECTED

            async def search_creator(self, name):
                if not type(self).succeed:
                    return LookupResult(
                        normalized_username=name,
                        status=RowStatus.FAILED,
                        error_code=ErrorCode.CREATOR_NOT_FOUND,
                    )
                return LookupResult(
                    normalized_username=name,
                    gmv=parse_gmv("$5K"),
                    status=RowStatus.SUCCESS,
                )

            async def close(self):
                return None

        app = create_app(JobStore(str(tmp_path / "retry-jobs")), RetrySession)
        async with AsyncClient(
            transport=ASGITransport(app=app), base_url="http://test"
        ) as client:
            response = await client.post(
                "/jobs",
                files={"file": ("retry.xlsx", _xlsx_bytes(), "application/xlsx")},
            )
            job_id = response.json()["job_id"]
            await _poll(client, job_id, {"completed_with_errors"})
            RetrySession.succeed = True
            assert (await client.post(f"/jobs/{job_id}/retry")).status_code == 200
            await _poll(client, job_id, {"completed"})
            assert len(RetrySession.paths) == 2
            assert RetrySession.paths[0] != RetrySession.paths[1]

        class RestartSession:
            paths = []
            instances = 0

            def __init__(self, profile):
                self.profile = profile
                type(self).instances += 1
                self.number = type(self).instances
                type(self).paths.append(Path(profile.storage_root))

            async def start(self, concurrency=1):
                return ProfileStatus.CONNECTED

            async def search_creator(self, name):
                if self.number == 1:
                    return LookupResult(
                        normalized_username=name,
                        status=RowStatus.FAILED,
                        error_code=ErrorCode.BROWSER_ERROR,
                    )
                return LookupResult(
                    normalized_username=name,
                    gmv=parse_gmv("$5K"),
                    status=RowStatus.SUCCESS,
                )

            async def close(self):
                return None

        restart_app = create_app(
            JobStore(str(tmp_path / "restart-jobs")), RestartSession
        )
        async with AsyncClient(
            transport=ASGITransport(app=restart_app), base_url="http://test"
        ) as client:
            response = await client.post(
                "/jobs",
                files={"file": ("restart.xlsx", _xlsx_bytes(), "application/xlsx")},
            )
            await _poll(client, response.json()["job_id"], {"completed"})
            assert RestartSession.instances == 2
            assert len(set(RestartSession.paths)) == 1
            await _wait_paths_removed(RestartSession.paths)

    monkeypatch.setattr("gmv.job_runner.MAX_SESSION_RESTARTS_PER_CREATOR", 2)
    asyncio.run(go())
