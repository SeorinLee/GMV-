"""API tests for independent browser and market selection."""

import asyncio
import io
from pathlib import Path

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook

from gmv.api import _job_payload_with_uploaded_rows, create_app
from gmv.config import ProfileStatus
from gmv.gmv_parser import parse_gmv
from gmv.job_runner import run_job
from gmv.models import ErrorCode, GmvValueType, LookupResult, RowStatus
from gmv.service import process_job
from gmv.store import JobRecord, JobStatus, JobStore


@pytest.fixture(autouse=True)
def _isolated_browser_profiles(tmp_path, monkeypatch):
    """API fakes still exercise clone lifecycle without reading the user's real login profile."""
    monkeypatch.setenv("GMV_PROFILE_ROOT", str(tmp_path / "profiles"))
    monkeypatch.setenv("GMV_RUNTIME_PROFILE_ROOT", str(tmp_path / "runtime_profiles"))


class FakeSession:
    def __init__(self, profile, gmv_map):
        self.profile = profile
        self.gmv_map = gmv_map

    async def start(self):
        return ProfileStatus.CONNECTED

    async def close(self):
        return None

    async def search_creator(self, name):
        raw = self.gmv_map.get(name)
        if raw is None:
            return LookupResult(
                normalized_username=name,
                status=RowStatus.FAILED,
                error_code=ErrorCode.CREATOR_NOT_FOUND,
            )
        parsed = parse_gmv(raw)
        status = (
            RowStatus.RANGE
            if parsed.value_type in (GmvValueType.RANGE_MAX, GmvValueType.OPEN_ENDED_ESTIMATE)
            else RowStatus.SUCCESS
        )
        return LookupResult(normalized_username=name, gmv=parsed, status=status, items_sold=7)


def _xlsx_bytes():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Creator Name"
    ws["A2"] = "@alpha"
    ws["A3"] = "beta"
    ws["A4"] = "gamma"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make(tmp_path, gmv_map):
    store = JobStore(str(tmp_path / "jobs"))
    app = create_app(store=store, session_factory=lambda p: FakeSession(p, gmv_map))
    client = AsyncClient(transport=ASGITransport(app=app), base_url="http://t")
    return app, client


async def _poll(client, job_id, terminal, tries=300):
    last = None
    for _ in range(tries):
        last = (await client.get(f"/jobs/{job_id}")).json()
        if last["status"] in terminal:
            return last
        await asyncio.sleep(0.02)
    raise AssertionError(f"never reached {terminal}: last={last['status'] if last else None}")


def test_profiles_endpoint(tmp_path):
    async def go():
        _, client = _make(tmp_path, {})
        async with client:
            codes = {p["profile_code"] for p in (await client.get("/profiles")).json()}
            assert codes == {"US_CHROME", "UK_CHROME", "US_EDGE", "UK_EDGE"}

    asyncio.run(go())


def test_job_payload_lists_uploaded_creators_before_first_lookup(tmp_path):
    store = JobStore(str(tmp_path / "jobs"))
    record = JobRecord(original_filename="manager.xlsx", status=JobStatus.QUEUED)
    store.save(record)
    store.input_path(record.id).write_bytes(_xlsx_bytes())

    payload = _job_payload_with_uploaded_rows(store, record)

    assert [row["creator"] for row in payload["rows"]] == ["@alpha", "beta", "gamma"]
    assert all(row["gmv_value"] is None for row in payload["rows"])
    assert all(row["items_sold"] is None for row in payload["rows"])


def test_profile_alias_returns_sanitized_payload(tmp_path):
    async def go():
        _, client = _make(tmp_path, {})
        async with client:
            payload = (await client.get("/profile")).json()
            assert payload["profile_code"] == "US_CHROME"
            assert "browser_channel" not in payload
            assert "affiliate_url" not in payload

    asyncio.run(go())


def test_health_exposes_worker_build_for_safe_launcher_restart(tmp_path):
    async def go():
        _, client = _make(tmp_path, {})
        async with client:
            payload = (await client.get("/health")).json()
            assert payload == {
                "ok": True,
                "build": "invitation-creators-v25",
                "mode": "normal",
            }

    asyncio.run(go())


def test_jobs_accept_without_profile_code_and_use_default(tmp_path):
    async def go():
        app, client = _make(tmp_path, {"alpha": "$5K", "beta": "$1K-$5K"})
        async with client:
            resp = await client.post(
                "/jobs",
                files={"file": ("manager.xlsx", _xlsx_bytes(), "application/xlsx")},
            )
            assert resp.status_code == 200
            payload = resp.json()
            job = app.state.store.get(payload["job_id"])
            assert job is not None
            assert job.selected_profile_code == "US_CHROME"
            assert job.concurrency == 1

    asyncio.run(go())


def test_jobs_accept_selected_uk_edge_profile(tmp_path):
    async def go():
        app, client = _make(tmp_path, {"alpha": "$5K"})
        async with client:
            resp = await client.post(
                "/jobs",
                data={"profile_code": "UK_EDGE"},
                files={"file": ("manager.xlsx", _xlsx_bytes(), "application/xlsx")},
            )
            assert resp.status_code == 200
            job = app.state.store.get(resp.json()["job_id"])
            assert job is not None
            assert job.selected_profile_code == "UK_EDGE"

    asyncio.run(go())


def test_jobs_reject_unknown_profile(tmp_path):
    async def go():
        _, client = _make(tmp_path, {})
        async with client:
            resp = await client.post(
                "/jobs",
                data={"profile_code": "UNKNOWN"},
                files={"file": ("manager.xlsx", _xlsx_bytes(), "application/xlsx")},
            )
            assert resp.status_code == 400
            assert resp.json()["code"] == "INVALID_PROFILE"

    asyncio.run(go())


def test_full_flow_and_retry(tmp_path):
    async def go():
        app, client = _make(tmp_path, {"alpha": "$5K", "beta": "$1K-$5K"})
        async with client:
            resp = await client.post(
                "/jobs",
                files={"file": ("manager.xlsx", _xlsx_bytes(), "application/xlsx")},
            )
            assert resp.status_code == 200
            job_id = resp.json()["job_id"]
            assert resp.json()["unique_creators"] == 3

            done = await _poll(client, job_id, {"completed", "completed_with_errors"})
            assert done["status"] == "completed_with_errors"
            assert (done["success"], done["range_rows"], done["failed"]) == (1, 1, 1)

            dl = await client.get(f"/jobs/{job_id}/download")
            assert dl.status_code == 200
            ws = load_workbook(io.BytesIO(dl.content)).active
            headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
            assert ws.cell(row=2, column=headers["GMV"]).value == 5000

            app.state.session_factory = lambda p: FakeSession(
                p, {"alpha": "$5K", "beta": "$1K-$5K", "gamma": "$2M"}
            )
            await client.post(f"/jobs/{job_id}/retry")
            retried = await _poll(client, job_id, {"completed"})
            assert retried["failed"] == 0

            dl2 = await client.get(f"/jobs/{job_id}/download")
            ws2 = load_workbook(io.BytesIO(dl2.content)).active
            assert ws2.cell(row=4, column=headers["GMV"]).value == 2_000_000
            assert ws2.cell(row=2, column=headers["GMV"]).value == 5000

    asyncio.run(go())


def test_bad_upload_rejected(tmp_path):
    async def go():
        _, client = _make(tmp_path, {})
        async with client:
            resp = await client.post(
                "/jobs",
                files={"file": ("bad.xlsx", b"not an excel file", "application/xlsx")},
            )
            assert resp.status_code == 400

    asyncio.run(go())


def test_invalid_file_type_returns_expected_code(tmp_path):
    async def go():
        _, client = _make(tmp_path, {})
        async with client:
            resp = await client.post(
                "/jobs",
                files={"file": ("bad.txt", b"not an excel file", "text/plain")},
            )
            assert resp.status_code == 400
            assert resp.json()["code"] == "INVALID_FILE_TYPE"

    asyncio.run(go())


def test_source_mutation_reservation_does_not_reject_job(tmp_path):
    async def go():
        app, client = _make(tmp_path, {})
        async with client:
            app.state.busy_profiles.add("US_CHROME")
            try:
                resp = await client.post(
                    "/jobs",
                    files={"file": ("manager.xlsx", _xlsx_bytes(), "application/xlsx")},
                )
                assert resp.status_code == 200
                await _poll(
                    client,
                    resp.json()["job_id"],
                    {"completed", "completed_with_errors"},
                )
            finally:
                app.state.busy_profiles.clear()

    asyncio.run(go())


def test_edge_browser_does_not_block_chrome_job(tmp_path):
    async def go():
        app, client = _make(tmp_path, {"alpha": "$5K"})
        async with client:
            app.state.busy_profiles.add("UK_EDGE")
            try:
                resp = await client.post(
                    "/jobs",
                    data={"profile_code": "US_CHROME"},
                    files={"file": ("manager.xlsx", _xlsx_bytes(), "application/xlsx")},
                )
                assert resp.status_code == 200
                await _poll(client, resp.json()["job_id"], {"completed", "completed_with_errors"})
            finally:
                app.state.busy_profiles.clear()

    asyncio.run(go())


def test_login_blocked_while_job_running(tmp_path):
    async def go():
        app, client = _make(tmp_path, {})
        async with client:
            app.state.busy_profiles.add("US_CHROME")
            try:
                resp = await client.post("/profile/login")
                assert resp.status_code == 409
                assert resp.json()["code"] == "BROWSER_BUSY"
            finally:
                app.state.busy_profiles.clear()

    asyncio.run(go())


def test_verify_blocked_while_job_running(tmp_path):
    async def go():
        app, client = _make(tmp_path, {})
        async with client:
            app.state.busy_profiles.add("US_CHROME")
            try:
                resp = await client.post("/profile/verify")
                assert resp.status_code == 409
                assert resp.json()["code"] == "BROWSER_BUSY"
            finally:
                app.state.busy_profiles.clear()

    asyncio.run(go())


def test_creator_column_missing_returns_code(tmp_path):
    async def go():
        _, client = _make(tmp_path, {})
        wb = Workbook()
        wb.active["A1"] = "Amount"
        buf = io.BytesIO()
        wb.save(buf)
        async with client:
            resp = await client.post(
                "/jobs",
                files={"file": ("nocreator.xlsx", buf.getvalue(), "application/xlsx")},
            )
            assert resp.status_code == 400
            assert resp.json()["code"] == "CREATOR_COLUMN_NOT_FOUND"

    asyncio.run(go())


def test_empty_template_returns_no_creators(tmp_path):
    async def go():
        _, client = _make(tmp_path, {})
        wb = Workbook()
        wb.active["A1"] = "Creator Name"
        buf = io.BytesIO()
        wb.save(buf)
        async with client:
            resp = await client.post(
                "/jobs",
                files={"file": ("empty.xlsx", buf.getvalue(), "application/xlsx")},
            )
            assert resp.status_code == 400
            assert resp.json()["code"] == "NO_CREATORS"

    asyncio.run(go())


def test_failed_upload_is_cleaned_up(tmp_path):
    async def go():
        app, client = _make(tmp_path, {})
        async with client:
            resp = await client.post(
                "/jobs",
                files={"file": ("bad.xlsx", b"not really excel", "application/xlsx")},
            )
            assert resp.status_code == 400
            # No lingering job directory or profile reservation after cleanup (spec §13, §12).
            assert app.state.store.list() == []
            assert app.state.busy_profiles == set()

    asyncio.run(go())


def test_completed_job_releases_busy_and_cancel_event(tmp_path):
    async def go():
        app, client = _make(tmp_path, {"alpha": "$5K", "beta": "$1K-$5K"})
        async with client:
            resp = await client.post(
                "/jobs",
                files={"file": ("manager.xlsx", _xlsx_bytes(), "application/xlsx")},
            )
            job_id = resp.json()["job_id"]
            await _poll(client, job_id, {"completed", "completed_with_errors"})
            # Busy released and the cancel signal cleaned up after the job ends (spec §8, §12).
            assert app.state.busy_profiles == set()
            assert job_id not in app.state.job_cancel_events

    asyncio.run(go())


def test_cancel_nonexistent_job_returns_404(tmp_path):
    async def go():
        _, client = _make(tmp_path, {})
        async with client:
            resp = await client.post("/jobs/does-not-exist/cancel")
            assert resp.status_code == 404
            assert resp.json()["code"] == "JOB_NOT_FOUND"

    asyncio.run(go())


def test_cancel_finished_job_returns_409(tmp_path):
    async def go():
        app, client = _make(tmp_path, {})
        async with client:
            rec = JobRecord(original_filename="x.xlsx", status=JobStatus.COMPLETED)
            app.state.store.save(rec)
            resp = await client.post(f"/jobs/{rec.id}/cancel")
            assert resp.status_code == 409
            assert resp.json()["code"] == "JOB_ALREADY_FINISHED"

    asyncio.run(go())


def test_cancel_running_job_sets_cancel_requested(tmp_path):
    async def go():
        app, client = _make(tmp_path, {})
        async with client:
            rec = JobRecord(original_filename="x.xlsx", status=JobStatus.RUNNING)
            app.state.store.save(rec)
            event = asyncio.Event()
            app.state.job_cancel_events[rec.id] = event
            resp = await client.post(f"/jobs/{rec.id}/cancel")
            assert resp.status_code == 200
            body = resp.json()
            assert body["status"] == JobStatus.CANCEL_REQUESTED
            assert body["message"]
            assert event.is_set()  # the cancel signal reached the loop
            assert app.state.store.get(rec.id).status == JobStatus.CANCEL_REQUESTED

    asyncio.run(go())


class _CancellingSession:
    """A fake session that requests cancel right after the first creator is processed."""

    def __init__(self, profile, event):
        self.profile = profile
        self._event = event
        self.cancel_event = None

    async def start(self):
        return ProfileStatus.CONNECTED

    async def close(self):
        return None

    async def search_creator(self, name):
        self._event.set()  # user hits "작동 중지" after this one finishes
        return LookupResult(
            normalized_username=name,
            gmv=parse_gmv("$5K"),
            status=RowStatus.SUCCESS,
            items_sold=1,
        )


def _three_creator_xlsx(path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Creator Name"
    ws["A2"], ws["A3"], ws["A4"] = "alpha", "beta", "gamma"
    wb.save(path)


def test_cancel_keeps_done_rows_and_marks_rest_cancelled(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    _three_creator_xlsx(inp)
    event = asyncio.Event()

    result = asyncio.run(
        run_job(
            str(inp),
            str(out),
            session_factory=lambda p: _CancellingSession(p, event),
            cancel_event=event,
        )
    )

    assert result.cancelled is True
    # The first creator is kept; everything not yet processed becomes CANCELLED (spec §7).
    assert result.results["alpha"].status is RowStatus.SUCCESS
    assert result.results["beta"].status is RowStatus.CANCELLED
    assert result.results["gamma"].status is RowStatus.CANCELLED
    # A partial workbook is still written.
    assert Path(out).exists()


def test_process_job_cancel_sets_status_and_partial_filename(tmp_path):
    async def go():
        store = JobStore(str(tmp_path / "jobs"))
        rec = JobRecord(original_filename="mydata.xlsx", status=JobStatus.QUEUED)
        store.save(rec)
        _three_creator_xlsx(store.input_path(rec.id))
        rec.total_rows, rec.unique_creators = 3, 3
        store.save(rec)

        event = asyncio.Event()
        updated = await process_job(
            store,
            lambda p: _CancellingSession(p, event),
            rec.id,
            cancel_event=event,
        )

        assert updated.status == JobStatus.CANCELLED
        assert updated.output_filename is not None
        assert "부분완료" in updated.output_filename
        # cancelled rows recorded; already-done row kept.
        by_creator = {r.creator: r.status for r in updated.rows}
        assert by_creator["alpha"] == RowStatus.SUCCESS.value
        assert by_creator["beta"] == RowStatus.CANCELLED.value
        assert store.output_path(rec.id).exists()  # partial download available (spec §7, §10)

    asyncio.run(go())


def test_retry_after_cancel_resumes_only_incomplete_rows(tmp_path):
    async def go():
        store = JobStore(str(tmp_path / "jobs"))
        rec = JobRecord(original_filename="resume.xlsx", status=JobStatus.QUEUED)
        store.save(rec)
        _three_creator_xlsx(store.input_path(rec.id))
        rec.total_rows, rec.unique_creators = 3, 3
        store.save(rec)

        # First run stops after alpha. beta/gamma are persisted as CANCELLED.
        event = asyncio.Event()
        cancelled = await process_job(
            store,
            lambda p: _CancellingSession(p, event),
            rec.id,
            cancel_event=event,
        )
        assert cancelled.status == JobStatus.CANCELLED
        assert cancelled.success == 1

        # Resume uses the normal retry endpoint: successful alpha is preserved, while the two
        # cancelled creators are promoted to retryable rows and searched automatically.
        app = create_app(
            store=store,
            session_factory=lambda p: FakeSession(
                p, {"alpha": "$5K", "beta": "$6K", "gamma": "$7K"}
            ),
        )
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://t") as client:
            response = await client.post(f"/jobs/{rec.id}/retry")
            assert response.status_code == 200
            assert response.json()["resumed_rows"] == 2

            done = await _poll(client, rec.id, {"completed"})
            assert done["status"] == "completed"
            assert done["success"] == 3
            assert done["failed"] == 0

            by_creator = {row["creator"]: row for row in done["rows"]}
            assert by_creator["alpha"]["gmv_value"] == 5000
            assert by_creator["beta"]["gmv_value"] == 6000
            assert by_creator["gamma"]["gmv_value"] == 7000

    asyncio.run(go())


@pytest.mark.parametrize("profile_code", ["UK_CHROME", "US_EDGE"])
def test_jobs_accept_cross_browser_market_profiles(tmp_path, profile_code):
    async def go():
        app, client = _make(tmp_path, {"alpha": "$5K"})
        async with client:
            resp = await client.post(
                "/jobs",
                data={"profile_code": profile_code},
                files={"file": ("manager.xlsx", _xlsx_bytes(), "application/xlsx")},
            )
            assert resp.status_code == 200
            job = app.state.store.get(resp.json()["job_id"])
            assert job is not None
            assert job.selected_profile_code == profile_code

    asyncio.run(go())


def test_create_job_always_uses_one_browser_page(tmp_path, monkeypatch):
    async def go():
        monkeypatch.setenv("GMV_DEFAULT_CONCURRENCY", "5")
        monkeypatch.setenv("GMV_MAX_CONCURRENCY", "6")

        for supplied in (6, 99, None):
            app, client = _make(tmp_path / str(supplied), {})
            async with client:
                data = {} if supplied is None else {"concurrency": str(supplied)}
                response = await client.post(
                    "/jobs",
                    data=data,
                    files={"file": ("manager.xlsx", _xlsx_bytes(), "application/xlsx")},
                )
                assert response.status_code == 200
                record = app.state.store.get(response.json()["job_id"])
                assert record.concurrency == 1
                await _poll(client, record.id, {"completed", "completed_with_errors"})

    asyncio.run(go())


def test_retry_legacy_interrupted_job_without_rows_runs_full_input(tmp_path):
    async def go():
        store = JobStore(str(tmp_path / "jobs"))
        rec = JobRecord(
            original_filename="legacy-interrupted.xlsx",
            selected_profile_code="US_CHROME",
            status=JobStatus.FAILED,
            total_rows=3,
            unique_creators=3,
            processed=1,
            success=1,
        )
        store.save(rec)
        _three_creator_xlsx(store.input_path(rec.id))

        app = create_app(
            store=store,
            session_factory=lambda p: FakeSession(
                p, {"alpha": "$5K", "beta": "$6K", "gamma": "$7K"}
            ),
        )
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://t") as client:
            response = await client.post(f"/jobs/{rec.id}/retry")
            assert response.status_code == 200
            assert response.json()["full_restart"] is True
            assert response.json()["resumed_rows"] == 3

            done = await _poll(client, rec.id, {"completed"})
            assert done["success"] == 3
            assert done["failed"] == 0
            assert {row["gmv_value"] for row in done["rows"]} == {5000, 6000, 7000}

    asyncio.run(go())
