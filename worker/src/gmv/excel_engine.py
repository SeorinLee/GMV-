"""Excel input reader and minimal result-workbook writer.

The uploaded workbook is input-only.  The downloadable workbook is created separately and
contains one consolidated ``Results`` sheet with exactly three columns:
``Creator Name``, ``GMV`` and ``Items sold``. GMV and Items sold are stored as expanded numeric
values (for example ``$148.4K`` becomes ``148400`` and ``4.7K`` becomes ``4700``). Failed
lookups keep the two result cells blank; their diagnostics remain available on the optional
``Errors`` sheet.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from gmv.config import ACCOUNT_COLUMN_ALIASES, resolve_profile_code
from gmv.creator import normalize_username
from gmv.models import LookupResult, RowStatus

# Spec §2 creator-column aliases, pre-normalized (lowercase, no spaces/underscores).
_CREATOR_ALIASES = {
    "creatorname",
    "creator",
    "creatorusername",
    "username",
    "tiktokusername",
    "tiktokid",
    "handle",
    "크리에이터",
    "크리에이터명",
    "크리에이터이름",
    "틱톡아이디",
}

CREATOR_HEADER = "Creator Name"
GMV_HEADER = "GMV"
ITEMS_HEADER = "Items sold"

# The result sheet always uses these columns in this exact order.
DEFAULT_OUTPUT_HEADERS = [CREATOR_HEADER, GMV_HEADER, ITEMS_HEADER]

# Optional debug sheet header — deliberately NOT a creator alias so a re-uploaded output file
# never treats it as a processing sheet.
ERRORS_SHEET = "Errors"
ERRORS_HEADERS = ["조회 대상", "조회 상태", "오류 코드", "오류 메시지", "단계"]

HEADER_ROW = 1


def _norm_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "").replace("_", "")


@dataclass
class CreatorRow:
    sheet_name: str
    excel_row_number: int  # 1-based, as shown in Excel
    original_creator_value: str
    normalized_username: str | None
    account_code: str | None = None  # per-row profile from a Market/Region column (spec §10)


@dataclass
class SheetPlan:
    sheet_name: str
    creator_col: int  # 1-based
    gmv_col: int | None
    rows: list[CreatorRow] = field(default_factory=list)


@dataclass
class WorkbookPlan:
    input_path: str
    sheets: list[SheetPlan] = field(default_factory=list)

    @property
    def total_rows(self) -> int:
        return sum(len(s.rows) for s in self.sheets)

    def unique_usernames(self) -> list[str]:
        """Deduplicated normalized usernames across all sheets (spec §2, §11)."""
        seen: dict[str, None] = {}
        for sheet in self.sheets:
            for row in sheet.rows:
                if row.normalized_username:
                    seen.setdefault(row.normalized_username, None)
        return list(seen.keys())


def detect_creator_column(ws: Worksheet) -> int | None:
    """Return the 1-based column index of the creator column, or None (spec §2)."""
    for cell in ws[HEADER_ROW]:
        if _norm_header(cell.value) in _CREATOR_ALIASES:
            return cell.column
    return None


def _find_header(ws: Worksheet, header: str) -> int | None:
    target = _norm_header(header)
    for cell in ws[HEADER_ROW]:
        if _norm_header(cell.value) == target:
            return cell.column
    return None


def _find_account_column(ws: Worksheet) -> int | None:
    """1-based index of a Market/Region/Account column, or None (spec §10)."""
    for cell in ws[HEADER_ROW]:
        if _norm_header(cell.value) in ACCOUNT_COLUMN_ALIASES:
            return cell.column
    return None


def detect_sheets(path: str) -> list[str]:
    """Names of sheets that contain a creator column, in original order (spec §2)."""
    wb = load_workbook(path, read_only=True)
    try:
        return [ws.title for ws in wb.worksheets if detect_creator_column(ws) is not None]
    finally:
        wb.close()


def read_workbook(path: str, sheets: list[str] | None = None) -> WorkbookPlan:
    """Build a processing plan, preserving sheet order (spec §2)."""
    wb = load_workbook(path)
    try:
        plan = WorkbookPlan(input_path=path)
        for ws in wb.worksheets:
            if sheets is not None and ws.title not in sheets:
                continue
            creator_col = detect_creator_column(ws)
            if creator_col is None:
                continue
            gmv_col = _find_header(ws, GMV_HEADER)
            account_col = _find_account_column(ws)
            sheet_plan = SheetPlan(sheet_name=ws.title, creator_col=creator_col, gmv_col=gmv_col)
            for r in range(HEADER_ROW + 1, ws.max_row + 1):
                raw = ws.cell(row=r, column=creator_col).value
                if raw is None or str(raw).strip() == "":
                    continue  # skip empty rows (spec §2)
                account_code = None
                if account_col is not None:
                    account_code = resolve_profile_code(ws.cell(row=r, column=account_col).value)
                sheet_plan.rows.append(
                    CreatorRow(
                        sheet_name=ws.title,
                        excel_row_number=r,
                        original_creator_value=str(raw),
                        normalized_username=normalize_username(raw),
                        account_code=account_code,
                    )
                )
            plan.sheets.append(sheet_plan)
        return plan
    finally:
        wb.close()


def write_results(
    input_path: str,
    plan: WorkbookPlan,
    results: dict[str, LookupResult],
    output_path: str,
    *,
    errors_sheet: bool = True,
) -> str:
    """Create a separate result workbook with only Creator Name, GMV and Items sold.

    ``input_path`` remains in the signature because callers pass it as part of the job contract,
    but its workbook layout is deliberately not copied to the output.
    """
    del input_path
    wb = Workbook()
    try:
        ws = wb.active
        ws.title = "Results"
        cols = {header: index for index, header in enumerate(DEFAULT_OUTPUT_HEADERS, start=1)}
        for header, column in cols.items():
            cell = ws.cell(row=HEADER_ROW, column=column, value=header)
            cell.font = Font(bold=True)

        failures: list[tuple[str, LookupResult]] = []
        output_row = HEADER_ROW + 1
        for sheet_plan in plan.sheets:
            for row in sheet_plan.rows:
                ws.cell(row=output_row, column=cols[CREATOR_HEADER], value=row.original_creator_value)
                result = results.get(row.normalized_username or "")
                if result is not None:
                    _apply_row(ws, output_row, cols, result)
                if result is not None and result.status is RowStatus.FAILED:
                    failures.append((row.original_creator_value, result))
                output_row += 1

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:C{max(HEADER_ROW, output_row - 1)}"
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18

        if errors_sheet and failures:
            _write_errors_sheet(wb, failures)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path
    finally:
        wb.close()


def _apply_row(
    ws: Worksheet,
    excel_row: int,
    cols: dict[str, int],
    result: LookupResult,
) -> None:
    gmv = result.gmv
    is_failure = result.status is RowStatus.FAILED or gmv is None
    if is_failure:
        return  # leave GMV / 판매개수 blank for failed creators (spec §5)

    # Store expanded numbers so Excel receives 148400 / 4700, not "$148.4K" / "4.7K".
    if gmv is not None and gmv.value is not None:
        numeric_gmv = int(gmv.value) if gmv.value == gmv.value.to_integral_value() else float(gmv.value)
        ws.cell(row=excel_row, column=cols[GMV_HEADER], value=numeric_gmv)

    if result.items_sold is not None:
        ws.cell(row=excel_row, column=cols[ITEMS_HEADER], value=result.items_sold)


def _write_errors_sheet(wb, failures: list[tuple[str, LookupResult]]) -> None:
    ws = wb[ERRORS_SHEET] if ERRORS_SHEET in wb.sheetnames else wb.create_sheet(ERRORS_SHEET)
    for idx, header in enumerate(ERRORS_HEADERS, start=1):
        ws.cell(row=HEADER_ROW, column=idx, value=header)
    for r, (creator, result) in enumerate(failures, start=HEADER_ROW + 1):
        ws.cell(row=r, column=1, value=creator)
        ws.cell(row=r, column=2, value=result.status.value)
        ws.cell(row=r, column=3, value=result.error_code.value if result.error_code else None)
        ws.cell(row=r, column=4, value=result.error_message)
        ws.cell(row=r, column=5, value=result.current_stage)


def make_output_filename(
    original_filename: str, when: datetime | None = None, *, partial: bool = False
) -> str:
    """`원본파일명_GMV완료_YYYYMMDD_HHmm.xlsx`, or `_GMV부분완료_` when cancelled (spec §3, §7)."""
    when = when or datetime.now()
    stem = Path(original_filename).stem
    tag = "GMV부분완료" if partial else "GMV완료"
    return f"{stem}_{tag}_{when:%Y%m%d_%H%M}.xlsx"
