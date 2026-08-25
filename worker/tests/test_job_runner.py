"""Job runner tests for the selected browser/market profile flow."""

import asyncio

import pytest
from openpyxl import Workbook, load_workbook

from gmv.config import ProfileStatus
from gmv.gmv_parser import parse_gmv
from gmv.job_runner import run_job
from gmv.models import ErrorCode, GmvValueType, LookupResult, RowStatus


class FakeSession:
    instances: list["FakeSession"] = []

    def __init__(self, profile, gmv_map, status=ProfileStatus.CONNECTED):
        self.profile = profile
        self.gmv_map = gmv_map
        self.status = status
        self.searched: list[str] = []
        FakeSession.instances.append(self)

    async def start(self):
        return self.status

    async def close(self):
        return None

    async def search_creator(self, name):
        self.searched.append(name)
        raw = self.gmv_map.get(name)
        if raw is None:
            return LookupResult(
                normalized_username=name,
                status=RowStatus.FAILED,
                error_code=ErrorCode.CREATOR_NOT_FOUND,
                account_label=self.profile.profile_code,
            )
        parsed = parse_gmv(raw)
        status = (
            RowStatus.RANGE
            if parsed.value_type in (GmvValueType.RANGE_MAX, GmvValueType.OPEN_ENDED_ESTIMATE)
            else RowStatus.SUCCESS
        )
        return LookupResult(
            normalized_username=name, gmv=parsed, status=status, account_label=self.profile.profile_code
        )


def _factory(gmv_map, status=ProfileStatus.CONNECTED):
    def make(profile):
        return FakeSession(profile, gmv_map, status)
    return make


def _wb_single(path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Creator Name"
    ws["A2"] = "@alpha"
    ws["A3"] = "beta"
    ws["A4"] = "alpha"
    wb.save(path)


def _wb_mixed(path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Creator Name"
    ws["B1"] = "Market"
    ws["A2"], ws["B2"] = "alpha", "US"
    ws["A3"], ws["B3"] = "beta", "UK"
    ws["A4"], ws["B4"] = "gamma", "미국"
    wb.save(path)


def _wb_names(path, names):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Creator Name"
    for row, name in enumerate(names, start=2):
        ws.cell(row=row, column=1, value=name)
    wb.save(path)


def setup_function():
    FakeSession.instances.clear()


def test_dedupe_searches_once(tmp_path):
    src = tmp_path / "s.xlsx"
    _wb_single(src)
    res = asyncio.run(
        run_job(
            str(src),
            str(tmp_path / "out.xlsx"),
            session_factory=_factory({"alpha": "$5K", "beta": "$1K-$5K"}),
            selected_profile_code="US_CHROME",
        )
    )
    assert len(FakeSession.instances) == 1
    assert sorted(FakeSession.instances[0].searched) == ["alpha", "beta"]
    ws = load_workbook(res.output_path).active
    assert ws["B2"].value == 5000
    assert ws["B4"].value == 5000


def test_single_profile_ignores_row_accounts(tmp_path):
    src = tmp_path / "m.xlsx"
    _wb_mixed(src)
    res = asyncio.run(
        run_job(
            str(src),
            str(tmp_path / "out.xlsx"),
            session_factory=_factory({"alpha": "$5K", "beta": "$9K", "gamma": "$2M"}),
        )
    )
    assert len(FakeSession.instances) == 1
    assert FakeSession.instances[0].profile.profile_code == "US_CHROME"
    assert res.progress.success == 3


@pytest.mark.parametrize(
    ("profile_code", "browser_channel", "market", "shop_region", "affiliate_host"),
    [
        ("UK_CHROME", "chrome", "UK", "GB", "https://affiliate.tiktok.com/"),
        ("US_EDGE", "msedge", "US", "US", "https://affiliate-us.tiktok.com/"),
    ],
)
def test_cross_browser_market_selection_reaches_automation_session(
    tmp_path, profile_code, browser_channel, market, shop_region, affiliate_host
):
    src = tmp_path / f"{profile_code}.xlsx"
    _wb_names(src, ["alpha"])

    asyncio.run(
        run_job(
            str(src),
            str(tmp_path / f"{profile_code}-out.xlsx"),
            session_factory=_factory({"alpha": "$5K"}),
            selected_profile_code=profile_code,
        )
    )

    profile = FakeSession.instances[0].profile
    assert profile.profile_code == profile_code
    assert profile.browser_channel == browser_channel
    assert profile.market == market
    assert profile.shop_region == shop_region
    assert profile.find_creators_url.startswith(affiliate_host)


def test_retry_only_requested(tmp_path):
    src = tmp_path / "r.xlsx"
    _wb_single(src)
    asyncio.run(
        run_job(
            str(src),
            str(tmp_path / "out.xlsx"),
            session_factory=_factory({"alpha": "$5K", "beta": "$1K"}),
            selected_profile_code="US_CHROME",
            only_usernames=["beta"],
        )
    )
    assert FakeSession.instances[0].searched == ["beta"]


def test_login_required_fails_all(tmp_path):
    src = tmp_path / "l.xlsx"
    _wb_single(src)
    res = asyncio.run(
        run_job(
            str(src),
            str(tmp_path / "out.xlsx"),
            session_factory=_factory({}, status=ProfileStatus.LOGIN_REQUIRED),
            selected_profile_code="US_CHROME",
        )
    )
    assert res.progress.failed == 2
    assert all(r.error_code is ErrorCode.LOGIN_REQUIRED for r in res.results.values())


def test_progress_callback_invoked(tmp_path):
    src = tmp_path / "p.xlsx"
    _wb_single(src)
    seen = []
    checkpoints = []

    async def cb(p):
        seen.append(p.processed)
        if p.last_username and p.last_result is not None:
            checkpoints.append(
                (p.last_username, p.last_result.status, p.last_result.gmv.value)
            )

    asyncio.run(
        run_job(
            str(src),
            str(tmp_path / "out.xlsx"),
            session_factory=_factory({"alpha": "$5K", "beta": "$1K"}),
            selected_profile_code="US_CHROME",
            progress_cb=cb,
        )
    )
    assert seen
    assert max(seen) == 2
    assert checkpoints[-2:] == [
        ("alpha", RowStatus.SUCCESS, 5000),
        ("beta", RowStatus.SUCCESS, 1000),
    ]


def test_missing_creator_is_failed_and_next_creator_is_processed(tmp_path):
    src = tmp_path / "missing-then-success.xlsx"
    _wb_single(src)
    result = asyncio.run(
        run_job(
            str(src),
            str(tmp_path / "missing-then-success-out.xlsx"),
            session_factory=_factory({"beta": "$7K"}),
            selected_profile_code="US_CHROME",
        )
    )

    assert FakeSession.instances[0].searched == ["alpha", "beta"]
    assert result.results["alpha"].status is RowStatus.FAILED
    assert result.results["alpha"].error_code is ErrorCode.CREATOR_NOT_FOUND
    assert result.results["beta"].status is RowStatus.SUCCESS


def test_watchdog_restarts_browser_and_retries_same_creator(tmp_path, monkeypatch):
    """A permanently pending Playwright call must not freeze the remaining workbook."""
    src = tmp_path / "watchdog.xlsx"
    _wb_single(src)

    class HangingOnceSession(FakeSession):
        created = 0

        def __init__(self, profile):
            type(self).created += 1
            self.number = type(self).created
            super().__init__(profile, {"alpha": "$5K", "beta": "$7K"})

        async def search_creator(self, name):
            self.searched.append(name)
            if name == "beta" and self.number == 1:
                await asyncio.Event().wait()
            parsed = parse_gmv("$5K" if name == "alpha" else "$7K")
            return LookupResult(
                normalized_username=name,
                gmv=parsed,
                status=RowStatus.SUCCESS,
                account_label=self.profile.profile_code,
            )

        async def capture_window_state(self):
            return "minimized"

    monkeypatch.setattr("gmv.job_runner.LOOKUP_WATCHDOG_SECONDS", 0.05)
    monkeypatch.setattr("gmv.job_runner.WATCHDOG_POLL_SECONDS", 0.005)
    monkeypatch.setattr("gmv.job_runner.SESSION_CLOSE_TIMEOUT_SECONDS", 0.05)
    monkeypatch.setattr("gmv.job_runner.MAX_SESSION_RESTARTS_PER_CREATOR", 2)

    result = asyncio.run(
        run_job(
            str(src),
            str(tmp_path / "watchdog-out.xlsx"),
            session_factory=HangingOnceSession,
            selected_profile_code="US_CHROME",
        )
    )

    assert HangingOnceSession.created == 2
    assert FakeSession.instances[-1].startup_window_state == "minimized"
    assert result.results["alpha"].status is RowStatus.SUCCESS
    assert result.results["beta"].status is RowStatus.SUCCESS


def test_watchdog_does_not_timeout_visible_security_puzzle(monkeypatch):
    from gmv.job_runner import _await_with_watchdog

    class PuzzleSession:
        security_challenge_active = True

    session = PuzzleSession()
    monkeypatch.setattr("gmv.job_runner.WATCHDOG_POLL_SECONDS", 0.005)

    async def go():
        async def clear_puzzle():
            await asyncio.sleep(0.06)
            session.security_challenge_active = False

        asyncio.create_task(clear_puzzle())
        return await _await_with_watchdog(
            asyncio.sleep(0.08),
            session=session,
            timeout_seconds=0.03,
        )

    asyncio.run(go())


def test_concurrent_lookup_overlaps_and_preserves_dedupe_and_progress(tmp_path):
    src = tmp_path / "parallel.xlsx"
    _wb_names(src, ["alpha", "beta", "alpha", "gamma", "delta"])
    checkpoints = []

    class OverlapSession(FakeSession):
        running = 0
        max_running = 0

        async def search_creator(self, name):
            type(self).running += 1
            type(self).max_running = max(type(self).max_running, type(self).running)
            try:
                await asyncio.sleep(0.02)
                return await super().search_creator(name)
            finally:
                type(self).running -= 1

    async def cb(progress):
        checkpoints.append(progress.processed)

    result = asyncio.run(
        run_job(
            str(src),
            str(tmp_path / "parallel-out.xlsx"),
            session_factory=lambda profile: OverlapSession(
                profile,
                {name: "$5K" for name in ("alpha", "beta", "gamma", "delta")},
            ),
            selected_profile_code="US_CHROME",
            concurrency=4,
            progress_cb=cb,
        )
    )

    assert OverlapSession.max_running > 1
    assert sorted(FakeSession.instances[0].searched) == ["alpha", "beta", "delta", "gamma"]
    assert result.progress.processed == result.progress.success == 4
    assert max(checkpoints) == 4


def test_concurrent_retry_searches_only_requested_names(tmp_path):
    src = tmp_path / "parallel-retry.xlsx"
    _wb_names(src, ["alpha", "beta", "gamma", "delta"])

    result = asyncio.run(
        run_job(
            str(src),
            str(tmp_path / "parallel-retry-out.xlsx"),
            session_factory=_factory({"beta": "$6K", "delta": "$8K"}),
            selected_profile_code="US_CHROME",
            concurrency=4,
            only_usernames=["beta", "delta"],
        )
    )

    assert sorted(FakeSession.instances[0].searched) == ["beta", "delta"]
    assert set(result.results) == {"beta", "delta"}
    assert result.progress.success == 2


def test_concurrent_cancel_finishes_inflight_and_marks_remaining_cancelled(tmp_path):
    src = tmp_path / "parallel-cancel.xlsx"
    out = tmp_path / "parallel-cancel-out.xlsx"
    names = [f"creator_{index}" for index in range(8)]
    _wb_names(src, names)
    cancel_event = asyncio.Event()

    class CancellingSession(FakeSession):
        async def search_creator(self, name):
            self.searched.append(name)
            await asyncio.sleep(0.01 if name == "creator_0" else 0.03)
            if name == "creator_0":
                cancel_event.set()
            parsed = parse_gmv("$5K")
            return LookupResult(
                normalized_username=name,
                gmv=parsed,
                status=RowStatus.SUCCESS,
                account_label=self.profile.profile_code,
            )

    result = asyncio.run(
        run_job(
            str(src),
            str(out),
            session_factory=lambda profile: CancellingSession(profile, {}),
            selected_profile_code="US_CHROME",
            concurrency=3,
            cancel_event=cancel_event,
        )
    )

    assert result.cancelled is True
    assert any(item.status is RowStatus.SUCCESS for item in result.results.values())
    assert any(item.status is RowStatus.CANCELLED for item in result.results.values())
    assert out.exists()
