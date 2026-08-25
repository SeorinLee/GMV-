"""Excel engine tests for the MVP workflow."""

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from gmv.excel_engine import (
    detect_creator_column,
    detect_sheets,
    make_output_filename,
    read_workbook,
    write_results,
)
from gmv.gmv_parser import parse_gmv
from gmv.models import ErrorCode, LookupResult, RowStatus
from gmv.service import UploadValidationError, validate_upload
from gmv.store import JobStore


def _styled_workbook(path, creator_header="Creator Name", with_gmv=True):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = creator_header
    ws["A1"].font = Font(bold=True)
    ws["A1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    col = 2
    if with_gmv:
        ws.cell(row=1, column=col, value="GMV")
        col += 1
    ws.cell(row=1, column=col, value="Note")

    ws["A2"] = "@alpha"
    ws["A3"] = "beta"
    ws["A4"] = "https://www.tiktok.com/@gamma"
    ws["A5"] = "alpha"
    if with_gmv:
        ws["B2"] = 100
        ws["B3"] = 999
    ws["C2"] = "=1+2"

    ws.column_dimensions["A"].width = 30
    ws.merge_cells("E1:F1")
    ws.freeze_panes = "A2"
    wb.save(path)
    return path


def test_detect_english_column(tmp_path):
    p = _styled_workbook(tmp_path / "en.xlsx")
    wb = load_workbook(p)
    assert detect_creator_column(wb["Data"]) == 1


@pytest.mark.parametrize("header", ["크리에이터", "크리에이터 이름", "TikTok Username", "handle", "USER_NAME"])
def test_detect_alias_and_korean_columns(tmp_path, header):
    p = _styled_workbook(tmp_path / "alias.xlsx", creator_header=header)
    wb = load_workbook(p)
    assert detect_creator_column(wb["Data"]) == 1


def test_multi_sheet_detection(tmp_path):
    p = tmp_path / "multi.xlsx"
    wb = Workbook()
    wb.active.title = "HasCreator"
    wb["HasCreator"]["A1"] = "Creator"
    wb["HasCreator"]["A2"] = "someone"
    wb.create_sheet("NoCreator")["A1"] = "Amount"
    wb.save(p)
    assert detect_sheets(str(p)) == ["HasCreator"]


def test_dedupe_unique_usernames(tmp_path):
    p = _styled_workbook(tmp_path / "dupe.xlsx")
    plan = read_workbook(str(p))
    assert plan.total_rows == 4
    assert plan.unique_usernames() == ["alpha", "beta", "gamma"]


def test_url_username_extracted_on_read(tmp_path):
    p = _styled_workbook(tmp_path / "url.xlsx")
    plan = read_workbook(str(p))
    normalized = {r.excel_row_number: r.normalized_username for r in plan.sheets[0].rows}
    assert normalized[4] == "gamma"


def _raw_results():
    return {
        "alpha": LookupResult(
            normalized_username="alpha",
            gmv=parse_gmv("$150.1K"),
            items_sold=4800,
            items_sold_raw="4.8K",
            status=RowStatus.SUCCESS,
        ),
        "beta": LookupResult(
            normalized_username="beta",
            status=RowStatus.FAILED,
            error_code=ErrorCode.CREATOR_NOT_FOUND,
            error_message="not found",
            current_stage="MATCH_USERNAME",
        ),
        "gamma": LookupResult(
            normalized_username="gamma",
            gmv=parse_gmv("$0-$5K"),
            items_sold=3,
            items_sold_raw="3",
            status=RowStatus.RANGE,
        ),
    }


def _headers(ws):
    return {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}


def test_output_has_only_gmv_and_items_columns(tmp_path):
    p = _styled_workbook(tmp_path / "cols.xlsx", with_gmv=False)
    plan = read_workbook(str(p))
    out = write_results(str(p), plan, _raw_results(), str(tmp_path / "out.xlsx"))
    wb = load_workbook(out)
    ws = wb["Results"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert headers == ["Creator Name", "GMV", "Items sold"]
    assert "Data" not in wb.sheetnames


def test_gmv_column_auto_created_with_raw_value(tmp_path):
    p = _styled_workbook(tmp_path / "nogmv.xlsx", with_gmv=False)
    plan = read_workbook(str(p))
    assert plan.sheets[0].gmv_col is None
    out = write_results(str(p), plan, _raw_results(), str(tmp_path / "out2.xlsx"))
    ws = load_workbook(out)["Results"]
    h = _headers(ws)
    assert "GMV" in h
    assert ws.cell(row=2, column=h["GMV"]).value == 150100
    assert ws.cell(row=4, column=h["GMV"]).value == 5000


def test_items_column_auto_created_with_raw_display(tmp_path):
    p = _styled_workbook(tmp_path / "items.xlsx", with_gmv=False)
    plan = read_workbook(str(p))
    out = write_results(str(p), plan, _raw_results(), str(tmp_path / "out.xlsx"))
    ws = load_workbook(out)["Results"]
    h = _headers(ws)
    assert "Items sold" in h
    assert ws.cell(row=2, column=h["Items sold"]).value == 4800


def test_failed_creator_gmv_and_items_blank(tmp_path):
    p = _styled_workbook(tmp_path / "fail.xlsx", with_gmv=False)
    plan = read_workbook(str(p))
    out = write_results(str(p), plan, _raw_results(), str(tmp_path / "out.xlsx"))
    ws = load_workbook(out)["Results"]
    h = _headers(ws)
    assert ws.cell(row=3, column=h["GMV"]).value is None  # beta failed
    assert ws.cell(row=3, column=h["Items sold"]).value is None


def test_errors_sheet_lists_failures_but_is_not_processed(tmp_path):
    p = _styled_workbook(tmp_path / "err.xlsx", with_gmv=False)
    plan = read_workbook(str(p))
    out = write_results(str(p), plan, _raw_results(), str(tmp_path / "out.xlsx"))
    wb = load_workbook(out)
    assert "Errors" in wb.sheetnames
    assert wb["Errors"].cell(row=1, column=1).value == "조회 대상"
    assert detect_sheets(out) == ["Results"]  # Errors is not a creator sheet


def test_creator_name_only_input_validates(tmp_path):
    store = JobStore(str(tmp_path / "jobs"))
    path = store.input_path("only")
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Creator Name"  # no GMV / 판매개수 / 조회 상태 / etc.
    ws["A2"] = "yourboybigmike"
    wb.save(path)
    total, unique = validate_upload(store, "only")  # must NOT raise
    assert (total, unique) == (1, 1)


def test_output_is_separate_from_input_format(tmp_path):
    p = _styled_workbook(tmp_path / "fmt.xlsx")
    plan = read_workbook(str(p))
    out = write_results(str(p), plan, _raw_results(), str(tmp_path / "out3.xlsx"))
    wb = load_workbook(out)
    ws = wb["Results"]
    assert ws["A1"].font.bold is True
    assert ws.column_dimensions["A"].width == 30
    assert ws.freeze_panes == "A2"
    assert [ws.cell(1, c).value for c in range(1, 4)] == [
        "Creator Name",
        "GMV",
        "Items sold",
    ]
    assert "Data" not in wb.sheetnames


def _template_path() -> Path:
    root = Path(__file__).resolve().parents[2]
    monorepo_path = root / "apps" / "web" / "public" / "gmv_upload_template.xlsx"
    portable_path = root / "web" / "public" / "gmv_upload_template.xlsx"
    return monorepo_path if monorepo_path.exists() else portable_path


def test_template_file_opens_and_detects_creator_column():
    template_path = _template_path()
    assert template_path.exists()
    wb = load_workbook(template_path)
    assert "Creators" in wb.sheetnames
    assert "사용방법" in wb.sheetnames
    assert detect_creator_column(wb["Creators"]) == 1


def test_template_only_creators_sheet_is_processed():
    # The instructions sheet must NOT be picked up as a processing target (spec §6).
    assert detect_sheets(str(_template_path())) == ["Creators"]


def test_template_has_only_creator_name_column():
    wb = load_workbook(_template_path())
    ws = wb["Creators"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert headers == ["Creator Name"]  # users only fill this; GMV/판매개수 are auto-added


def test_template_validation_reports_no_creators(tmp_path):
    store = JobStore(str(tmp_path / "jobs"))
    path = store.input_path("empty")
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Creator Name"
    wb.save(path)
    with pytest.raises(UploadValidationError) as exc:
        validate_upload(store, "empty")
    assert exc.value.code == "NO_CREATORS"


def test_missing_creator_column_reports_error(tmp_path):
    store = JobStore(str(tmp_path / "jobs"))
    path = store.input_path("missing")
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Amount"
    wb.save(path)
    with pytest.raises(UploadValidationError) as exc:
        validate_upload(store, "missing")
    assert exc.value.code == "CREATOR_COLUMN_NOT_FOUND"


def test_output_filename():
    name = make_output_filename("manager.xlsx", datetime(2026, 7, 14, 11, 30))
    assert name == "manager_GMV완료_20260714_1130.xlsx"
